from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl


url = "https://pje-consulta.tjce.jus.br/pje1grau/ConsultaPublica/listView.seam"


options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(15)
driver.get(url)
driver.fullscreen_window()

numero_oab = 29042

campo_oab = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)

drop_estados = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes = Select(drop_estados)
opcoes.select_by_visible_text("CE")

btn_pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
btn_pesquisar.click()

processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")

for processo in processos:
    processo.click()
    sleep(10)
    janelas = driver.window_handles     # ! verifica quais janelas estão disponíveis
    driver.switch_to.window(janelas[-1])
    driver.set_window_size(1920, 1080)
    numero_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12 ']")
    numero_processo = numero_processo[0]
    numero_processo = numero_processo.text
    
    data_distribuicao = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
    data_distribuicao = data_distribuicao[1]
    data_distribuicao = data_distribuicao.text
    
    movimentacoes = driver.find_elements(By.XPATH, "//div[@id='j_id131:processoEventoPanel_body']//tr[contains(@class,'rich-table-row')]//td//div//div//span")
    lista_movimentacoes = []
    
    for movimentacao in movimentacoes:
        lista_movimentacoes.append(movimentacao.text)
    
    workbook = openpyxl.load_workbook('/home/diego/Desktop/miseleniums/drafts/pje/dados.xlsx')
    try:
        pagina_processo = workbook[numero_processo]
        pagina_processo['A1'].value = "Número do Processo"
        pagina_processo['B1'].value = "Data da Distribuição"
        pagina_processo['C1'].value = "Movimentações"
        pagina_processo['A2'].value = numero_processo
        pagina_processo['B2'].value = data_distribuicao
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_rol=len(lista_movimentacoes), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save('/home/diego/Desktop/miseleniums/drafts/pje/dados.xlsx')
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
            

    except Exception as error:
        workbook.create_chartsheet(numero_processo)
        pagina_processo = workbook[numero_processo]
        pagina_processo['A1'].value = "Número do Processo"
        pagina_processo['B1'].value = "Data da Distribuição"
        pagina_processo['C1'].value = "Movimentações"
        pagina_processo['A2'].value = numero_processo
        pagina_processo['B2'].value = data_distribuicao
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_rol=len(lista_movimentacoes), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save('/home/diego/Desktop/miseleniums/drafts/pje/dados.xlsx')
        driver.close()
        sleep(3)
        driver.switch_to.window(driver.window_handles[0])